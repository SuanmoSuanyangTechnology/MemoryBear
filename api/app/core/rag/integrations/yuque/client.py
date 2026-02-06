"""Yuque API client for document operations."""

import os
import re
from typing import Optional, List
from datetime import datetime, timedelta
import httpx
import urllib.parse
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import zlib

from app.core.rag.integrations.yuque.exceptions import (
    YuqueAuthError,
    YuqueAPIError,
    YuqueNotFoundError,
    YuquePermissionError,
    YuqueRateLimitError,
    YuqueNetworkError,
)
from app.core.rag.integrations.yuque.models import YuqueDocInfo, YuqueRepoInfo
from app.core.rag.integrations.yuque.retry import with_retry


class YuqueAPIClient:
    """Yuque API client for document synchronization."""
    
    def __init__(
        self,
        user_id: str,
        token: str,
        api_base_url: str = "https://www.yuque.com/api/v2",
        timeout: int = 30,
        max_retries: int = 3
    ):
        """
        Initialize Yuque API client.
        
        Args:
            user_id: Yuque user ID or login name
            token: Yuque personal access token
            api_base_url: Yuque API base URL
            timeout: Request timeout in seconds
            max_retries: Maximum number of retries
        """
        self.user_id = user_id
        self.token = token
        self.api_base_url = api_base_url
        self.timeout = timeout
        self.max_retries = max_retries
        self._http_client: Optional[httpx.AsyncClient] = None
    
    async def __aenter__(self):
        """Async context manager entry."""
        self._http_client = httpx.AsyncClient(
            base_url=self.api_base_url,
            timeout=self.timeout,
            headers={
                "Content-Type": "application/json",
                "X-Auth-Token": self.token,
                "User-Agent": "Yuque-Integration-Client"
            }
        )
        return self
    
    async def __aexit__(self, exc_type, exc_val, exc_tb):
        """Async context manager exit."""
        if self._http_client:
            await self._http_client.aclose()
    
    def _handle_api_error(self, response: httpx.Response):
        """Handle API error responses."""
        try:
            data = response.json()
        except Exception:
            data = {}
        
        status_code = response.status_code
        error_msg = data.get("message", "Unknown error")
        
        # Rate limit errors
        if status_code == 429:
            raise YuqueRateLimitError(
                f"Rate limit exceeded: {error_msg}",
                error_code=str(status_code),
                details=data
            )
        # Not found errors
        elif status_code == 404:
            raise YuqueNotFoundError(
                f"Resource not found: {error_msg}",
                error_code=str(status_code),
                details=data
            )
        # Permission errors
        elif status_code == 403:
            raise YuquePermissionError(
                f"Permission denied: {error_msg}",
                error_code=str(status_code),
                details=data
            )
        # Authentication errors
        elif status_code == 401:
            raise YuqueAuthError(
                f"Authentication failed: {error_msg}",
                error_code=str(status_code),
                details=data
            )
        # Generic API error
        else:
            raise YuqueAPIError(
                f"API error: {error_msg}",
                error_code=str(status_code),
                details=data
            )
    
    @with_retry
    async def get_user_repos(self) -> List[YuqueRepoInfo]:
        """
        Get all repositories (知识库) for the user.
        
        Returns:
            List of YuqueRepoInfo objects
            
        Raises:
            YuqueAPIError: If API call fails
        """
        try:
            if not self._http_client:
                raise YuqueAPIError("HTTP client not initialized")
            
            response = await self._http_client.get(f"/users/{self.user_id}/repos")
            
            if response.status_code != 200:
                self._handle_api_error(response)
            
            data = response.json()
            repos_data = data.get("data", [])
            
            repos = []
            for repo_data in repos_data:
                try:
                    repo = YuqueRepoInfo(
                        id=repo_data.get("id"),
                        type=repo_data.get("type", ""),
                        name=repo_data.get("name", ""),
                        namespace=repo_data.get("namespace", ""),
                        slug=repo_data.get("slug", ""),
                        description=repo_data.get("description"),
                        public=repo_data.get("public", 0),
                        items_count=repo_data.get("items_count", 0),
                        created_at=datetime.fromisoformat(repo_data.get("created_at", "").replace("Z", "+00:00")),
                        updated_at=datetime.fromisoformat(repo_data.get("updated_at", "").replace("Z", "+00:00"))
                    )
                    repos.append(repo)
                except (ValueError, TypeError, KeyError) as e:
                    # Skip invalid repo entries
                    continue
            
            return repos
            
        except httpx.HTTPError as e:
            raise YuqueAPIError(f"HTTP error: {str(e)}")
        except Exception as e:
            if isinstance(e, (YuqueAPIError, YuqueAuthError)):
                raise
            raise YuqueAPIError(f"Unexpected error: {str(e)}")
    
    @with_retry
    async def get_repo_docs(self, book_id: int) -> List[YuqueDocInfo]:
        """
        Get all documents in a repository.
        
        Args:
            book_id: repository id
            
        Returns:
            List of YuqueDocInfo objects (without body content)
            
        Raises:
            YuqueAPIError: If API call fails
        """
        try:
            if not self._http_client:
                raise YuqueAPIError("HTTP client not initialized")
            
            response = await self._http_client.get(f"/repos/{book_id}/docs")
            
            if response.status_code != 200:
                self._handle_api_error(response)
            
            data = response.json()
            docs_data = data.get("data", [])
            
            docs = []
            for doc_data in docs_data:
                try:
                    published_at = doc_data.get("published_at")
                    doc = YuqueDocInfo(
                        id=doc_data.get("id"),
                        type=doc_data.get("type", ""),
                        slug=doc_data.get("slug", ""),
                        title=doc_data.get("title", ""),
                        book_id=doc_data.get("book_id"),
                        format=doc_data.get("format", "markdown"),
                        body=None,  # Body not included in list API
                        body_draft=None,
                        body_html=None,
                        public=doc_data.get("public", 0),
                        status=doc_data.get("status", 0),
                        created_at=datetime.fromisoformat(doc_data.get("created_at", "").replace("Z", "+00:00")),
                        updated_at=datetime.fromisoformat(doc_data.get("updated_at", "").replace("Z", "+00:00")),
                        published_at=datetime.fromisoformat(published_at.replace("Z", "+00:00")) if published_at else None,
                        word_count=doc_data.get("word_count", 0),
                        cover=doc_data.get("cover"),
                        description=doc_data.get("description")
                    )
                    docs.append(doc)
                except (ValueError, TypeError, KeyError) as e:
                    # Skip invalid doc entries
                    continue
            
            return docs
            
        except httpx.HTTPError as e:
            raise YuqueAPIError(f"HTTP error: {str(e)}")
        except Exception as e:
            if isinstance(e, (YuqueAPIError, YuqueNotFoundError)):
                raise
            raise YuqueAPIError(f"Unexpected error: {str(e)}")
    
    @with_retry
    async def get_doc_detail(self, id: int) -> YuqueDocInfo:
        """
        Get detailed document information including content.
        
        Args:
            id: document ID
            
        Returns:
            YuqueDocInfo object with full content
            
        Raises:
            YuqueAPIError: If API call fails
        """
        try:
            if not self._http_client:
                raise YuqueAPIError("HTTP client not initialized")
            
            response = await self._http_client.get(
                f"/repos/docs/{id}",
                params={"raw": 1}  # Get raw markdown content
            )
            
            if response.status_code != 200:
                self._handle_api_error(response)
            
            data = response.json()
            doc_data = data.get("data", {})
            
            published_at = doc_data.get("published_at")
            doc = YuqueDocInfo(
                id=doc_data.get("id"),
                type=doc_data.get("type", ""),
                slug=doc_data.get("slug", ""),
                title=doc_data.get("title", ""),
                book_id=doc_data.get("book_id"),
                format=doc_data.get("format", "markdown"),
                body=doc_data.get("body", ""),
                body_draft=doc_data.get("body_draft"),
                body_html=doc_data.get("body_html"),
                public=doc_data.get("public", 0),
                status=doc_data.get("status", 0),
                created_at=datetime.fromisoformat(doc_data.get("created_at", "").replace("Z", "+00:00")),
                updated_at=datetime.fromisoformat(doc_data.get("updated_at", "").replace("Z", "+00:00")),
                published_at=datetime.fromisoformat(published_at.replace("Z", "+00:00")) if published_at else None,
                word_count=doc_data.get("word_count", 0),
                cover=doc_data.get("cover"),
                description=doc_data.get("description")
            )
            
            return doc
            
        except httpx.HTTPError as e:
            raise YuqueAPIError(f"HTTP error: {str(e)}")
        except Exception as e:
            if isinstance(e, (YuqueAPIError, YuqueNotFoundError)):
                raise
            raise YuqueAPIError(f"Unexpected error: {str(e)}")
    
    async def download_document(
        self,
        doc: YuqueDocInfo,
        save_dir: str
    ) -> str:
        """
        Download document content to local file.
        
        Args:
            doc: Document info (can be without body)
            save_dir: Directory to save the file
            
        Returns:
            Full path to the saved file
            
        Raises:
            YuqueAPIError: If download fails
        """
        try:
            # Get full document content if not already loaded
            if not doc.body:
                doc = await self.get_doc_detail(doc.id)
            
            # Sanitize filename
            filename = re.sub(r'[\/:*?"<>|]', '_', doc.title)

            # Determine file extension based on format
            content = doc.body or ""
            if doc.format == "markdown":
                file_extension = "md"
            elif doc.format == "lake":
                file_extension = "md"  # Save lake format as markdown
            elif doc.format == "html":
                file_extension = "html"
            elif doc.format == "lakesheet":
                file_extension = "xlsx"

                body_data = json.loads(doc.body)
                sheet_data = body_data.get("sheet", "")
                try:
                    sheet_raw = zlib.decompress(bytes(sheet_data, 'latin-1'))
                except Exception as e:
                    print(f"Error decompressing sheet data: {e}")
                    raise ValueError("Invalid or unsupported sheet data format.")
                try:
                    sheet_text = sheet_raw.decode("utf-8")  # 假设是 UTF-8 编码
                except UnicodeDecodeError:
                    sheet_text = sheet_raw.decode("gbk")  # 如果 UTF-8 解码失败，尝试 GBK

                file_full_path = os.path.join(save_dir, f"{filename}.{file_extension}")
                self.generate_excel_from_sheet(sheet_text, file_full_path)
                return file_full_path
            else:
                file_extension = "txt"

            file_full_path = os.path.join(save_dir, f"{filename}.{file_extension}")
            # Remove existing file if it exists
            if os.path.exists(file_full_path):
                os.remove(file_full_path)
            
            # Write content to file
            with open(file_full_path, "w", encoding="utf-8") as file:
                file.write(content)
            
            return file_full_path
            
        except Exception as e:
            if isinstance(e, YuqueAPIError):
                raise
            raise YuqueAPIError(f"Unexpected error during file download: {str(e)}")

    def generate_excel_from_sheet(self, sheet_text: str, save_path: str):
        """
        将解析的 sheet_text 数据转换为 Excel 文件。

        Args:
            sheet_text (str): JSON 格式的 sheet 数据。
            save_path (str): Excel 文件的保存路径。
        """
        try:
            # 解析 JSON 数据
            sheets = json.loads(sheet_text)

            if not isinstance(sheets, list):
                raise ValueError("sheet_text must be a JSON array of sheets.")

            # 创建一个新的 Excel 工作簿
            workbook = Workbook()

            for sheet_index, sheet_data in enumerate(sheets):
                sheet_name = sheet_data.get("name", f"Sheet{sheet_index + 1}")
                row_data = sheet_data.get("data", {})
                merge_cells = sheet_data.get("mergeCells", {})
                rows_styles = sheet_data.get("rows", [])
                cols_styles = sheet_data.get("columns", [])

                # 创建 Sheet
                if sheet_index == 0:
                    worksheet = workbook.active
                    worksheet.title = sheet_name
                else:
                    worksheet = workbook.create_sheet(title=sheet_name)

                # 设置列宽
                for col_index, col_style in enumerate(cols_styles):
                    col_width = col_style.get("size", 82.125) / 7.0
                    col_letter = get_column_letter(col_index + 1)  # Excel 列从1开始
                    worksheet.column_dimensions[col_letter].width = col_width

                # 设置行高
                for row_index, row_style in enumerate(rows_styles):
                    row_height = row_style.get("size", 24) / 1.5
                    worksheet.row_dimensions[row_index + 1].height = row_height

                # 写入单元格数据
                for r_index, row in row_data.items():
                    for c_index, cell in row.items():
                        # 防御性检查：确保行号和列号都是有效的整数
                        try:
                            row_number = int(r_index) + 1
                            col_number = int(c_index) + 1
                        except ValueError:
                            print(f"Invalid row or column index: r_index={r_index}, c_index={c_index}")
                            continue

                        if col_number < 1 or col_number > 16384:  # Excel 最大列数支持到 XFD，即 16384 列
                            print(f"Invalid column index: c_index={c_index}")
                            continue

                        cell_obj = worksheet.cell(row=row_number, column=col_number)

                        # 处理值和公式
                        cell_value = cell.get("value", "")
                        if isinstance(cell_value, dict):
                            # 检查是否为公式
                            if cell_value.get("class") == "formula" and "formula" in cell_value:
                                cell_obj.value = f"={cell_value['formula']}"  # 写入公式
                            else:
                                cell_obj.value = cell_value.get("value", "")  # 写入值
                        else:
                            cell_obj.value = cell_value  # 写入简单值

                        # 应用样式
                        style = cell.get("style", {})
                        self.apply_cell_style(cell_obj, style)

                # 合并单元格
                for key, merge_def in merge_cells.items():
                    start_row = merge_def["row"] + 1
                    start_col = merge_def["col"] + 1
                    end_row = start_row + merge_def["rowCount"] - 1
                    end_col = start_col + merge_def["colCount"] - 1
                    worksheet.merge_cells(
                        start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col
                    )

            # 保存 Excel 文件
            workbook.save(save_path)
            print(f"Excel file successfully saved to: {save_path}")

        except Exception as e:
            print(f"Error generating Excel file: {e}")


    def apply_cell_style(self, cell, style):
        """
        应用单元格样式，包括字体、对齐、背景颜色等。

        Args:
            cell: openpyxl 的单元格对象。
            style: 字典格式的样式信息。
        """
        # 定义允许的对齐值
        allowed_horizontal_alignments = {"general", "left", "center", "centerContinuous", "right", "fill", "justify",
                                         "distributed"}
        allowed_vertical_alignments = {"top", "center", "justify", "distributed", "bottom"}

        # 处理字体
        font = Font(
            size=style.get("fontSize", 11),
            bold=style.get("fontWeight", False),
            italic=style.get("fontStyle", "normal") == "italic",
            underline="single" if style.get("underline", False) else None,
            color=self.convert_color_to_hex(style.get("color", "#000000")),
        )
        cell.font = font

        # 处理对齐方式
        horizontal_alignment = style.get("hAlign", "left")
        vertical_alignment = style.get("vAlign", "top")

        # 如果对齐值无效，则使用默认值
        if horizontal_alignment not in allowed_horizontal_alignments:
            horizontal_alignment = "left"
        if vertical_alignment not in allowed_vertical_alignments:
            vertical_alignment = "top"

        alignment = Alignment(
            horizontal=horizontal_alignment,
            vertical=vertical_alignment,
            wrap_text=style.get("overflow") == "wrap",
        )
        cell.alignment = alignment

        # 处理背景颜色
        background_color = style.get("backColor", None)
        if background_color:
            hex_color = self.convert_color_to_hex(background_color)
            if hex_color:
                cell.fill = PatternFill(
                    start_color=hex_color,
                    end_color=hex_color,
                    fill_type="solid"
                )

    def convert_color_to_hex(self, color):
        """
        将颜色从 `rgba(...)` 或 `rgb(...)` 转换为 aRGB 十六进制格式。

        Args:
            color (str): 原始颜色字符串，如 `rgba(255,255,0,1.00)` 或 `#FFFFFF`。

        Returns:
            str: 转换后的颜色字符串（符合 openpyxl 的格式），例如 `FFFF0000`。
        """
        try:
            if not color:
                return None

            # 如果是 `#RRGGBB` 或 `#AARRGGBB` 格式，直接返回
            if color.startswith("#"):
                return color.lstrip("#").upper()

            # 如果是 `rgb(...)` 格式，例如 `rgb(255,255,0)`
            if color.startswith("rgb("):
                rgb_values = color.strip("rgb()").split(",")
                red, green, blue = [int(v) for v in rgb_values]
                return f"FF{red:02X}{green:02X}{blue:02X}"

            # 如果是 `rgba(...)` 格式，例如 `rgba(255,255,0,1.00)`
            if color.startswith("rgba("):
                rgba_values = color.strip("rgba()").split(",")
                red, green, blue = [int(v) for v in rgba_values[:3]]
                alpha = float(rgba_values[3])
                alpha_hex = int(alpha * 255)  # 将透明度转换为 [00, FF]
                return f"{alpha_hex:02X}{red:02X}{green:02X}{blue:02X}"

            # 返回默认颜色
            return None
        except Exception as e:
            print(f"Error parsing color '{color}': {e}")
            return None
