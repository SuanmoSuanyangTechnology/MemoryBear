import copy
import logging
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.core.rag.chunk.parser.base import DocumentParser
from app.core.rag.deepdoc.parser import ExcelParser as RAGExcelParser
from app.core.rag.nlp import add_positions, tokenize


EMPTY_MARKERS = {"", "/", "none", "null", "nan", "-"}


@dataclass
class _SheetMatrix:
    rows: list[tuple[int, list[str | None]]]
    start_col: int
    end_col: int


@dataclass
class _TableRegion:
    sheet_name: str
    subtable_id: int
    start_row: int
    end_row: int
    start_col: int
    end_col: int
    rows: list[tuple[int, list[str | None]]]


def _is_truthy(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    return str(value).strip().lower() == "true"


def _is_empty(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip().lower() in EMPTY_MARKERS
    return False


def _stringify(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float):
        if value != value:
            return None
        if value.is_integer():
            return str(int(value))
        return str(value)
    if isinstance(value, str):
        text = value.strip()
        return None if text.lower() in EMPTY_MARKERS else text
    return str(value).strip()


def _normalize_cell_value(cell) -> str | None:
    value = cell.value
    if isinstance(value, (int, float)) and "%" in str(getattr(cell, "number_format", "")):
        return f"{value * 100:g}%"
    return _stringify(value)


def _dedupe_headers(headers: list[str]) -> list[str]:
    counts: dict[str, int] = {}
    result = []
    for header in headers:
        counts[header] = counts.get(header, 0) + 1
        if counts[header] == 1:
            result.append(header)
        else:
            result.append(f"{header}_{counts[header]}")
    return result


def _col_range(start_col: int, end_col: int) -> str:
    return f"{get_column_letter(start_col)}:{get_column_letter(end_col)}"


def _cell_range(row_index: int, start_col: int, end_col: int) -> str:
    return f"{get_column_letter(start_col)}{row_index}:{get_column_letter(end_col)}{row_index}"


class ExcelParser(DocumentParser):
    def parse(self, ctx):
        binary = ctx.binary
        if not binary:
            with open(ctx.filename, "rb") as file:
                binary = file.read()

        excel_parser = RAGExcelParser()
        if _is_truthy(ctx.parser_config.get("html4excel")):
            return [(_, "") for _ in excel_parser.html(binary, 12) if _]
        return [(_, "") for _ in excel_parser(binary) if _]


class StructuredExcelParser(DocumentParser):
    def parse(self, ctx):
        binary = ctx.binary
        if not binary:
            with open(ctx.filename, "rb") as file:
                binary = file.read()

        workbook = self._load_workbook(binary)
        try:
            chunks = []
            for sheet in workbook.worksheets:
                matrix = self._sheet_matrix(sheet)
                for table in self._detect_tables(sheet.title, matrix):
                    chunks.extend(self._table_chunks(ctx, table))
            return chunks
        finally:
            close = getattr(workbook, "close", None)
            if callable(close):
                close()

    def _load_workbook(self, binary: bytes):
        stream = BytesIO(binary)
        try:
            return load_workbook(stream, data_only=True)
        except Exception as exc:
            logging.info(f"Structured Excel openpyxl load failed, using legacy workbook loader: {exc}")
            stream.seek(0)
            return RAGExcelParser._load_excel_to_workbook(stream)

    def _sheet_matrix(self, sheet) -> _SheetMatrix | None:
        max_row = sheet.max_row or 0
        max_col = sheet.max_column or 0
        if max_row == 0 or max_col == 0:
            return None

        values: list[list[str | None]] = []
        for row in range(1, max_row + 1):
            row_values = []
            for col in range(1, max_col + 1):
                row_values.append(_normalize_cell_value(sheet.cell(row=row, column=col)))
            values.append(row_values)

        for merged_range in getattr(sheet, "merged_cells", []).ranges:
            top_left = values[merged_range.min_row - 1][merged_range.min_col - 1]
            if _is_empty(top_left):
                continue
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    if _is_empty(values[row - 1][col - 1]):
                        values[row - 1][col - 1] = top_left

        non_empty_cols = [
            col_idx
            for col_idx in range(max_col)
            if any(not _is_empty(row[col_idx]) for row in values)
        ]
        if not non_empty_cols:
            return None

        start_offset = min(non_empty_cols)
        end_offset = max(non_empty_cols)
        rows = [
            (row_idx, row[start_offset:end_offset + 1])
            for row_idx, row in enumerate(values, start=1)
        ]
        return _SheetMatrix(
            rows=rows,
            start_col=start_offset + 1,
            end_col=end_offset + 1,
        )

    def _detect_tables(self, sheet_name: str, matrix: _SheetMatrix | None) -> list[_TableRegion]:
        if matrix is None:
            return []

        tables: list[_TableRegion] = []
        current_rows: list[tuple[int, list[str | None]]] = []

        def flush_current():
            nonlocal current_rows
            if len(current_rows) < 2:
                current_rows = []
                return

            non_empty_offsets = [
                col_idx
                for col_idx in range(len(current_rows[0][1]))
                if any(not _is_empty(row_values[col_idx]) for _, row_values in current_rows)
            ]
            if not non_empty_offsets:
                current_rows = []
                return

            start_offset = min(non_empty_offsets)
            end_offset = max(non_empty_offsets)
            trimmed_rows = [
                (row_idx, row_values[start_offset:end_offset + 1])
                for row_idx, row_values in current_rows
            ]
            tables.append(
                _TableRegion(
                    sheet_name=sheet_name,
                    subtable_id=len(tables) + 1,
                    start_row=current_rows[0][0],
                    end_row=current_rows[-1][0],
                    start_col=matrix.start_col + start_offset,
                    end_col=matrix.start_col + end_offset,
                    rows=trimmed_rows,
                )
            )
            current_rows = []

        for row_idx, row_values in matrix.rows:
            if all(_is_empty(value) for value in row_values):
                flush_current()
                continue
            current_rows.append((row_idx, row_values))

        flush_current()
        return tables

    def _table_chunks(self, ctx, table: _TableRegion) -> list[dict]:
        header_row_index, header_values = table.rows[0]
        del header_row_index
        headers = _dedupe_headers([
            value if not _is_empty(value) else f"列{get_column_letter(table.start_col + offset)}"
            for offset, value in enumerate(header_values)
        ])

        chunks = []
        for row_index, row_values in table.rows[1:]:
            fields = {
                header: value
                for header, value in zip(headers, row_values, strict=False)
                if not _is_empty(value)
            }
            if not fields:
                continue

            content = self._build_content(ctx.filename, table, row_index, fields)
            doc = copy.deepcopy(ctx.doc)
            add_positions(doc, [[row_index - 1, table.start_col, table.end_col, row_index, row_index]])
            tokenize(doc, content, ctx.is_english)
            doc["metadata"] = {
                "source_type": "excel",
                "source": {
                    "kind": "excel_table_row",
                    "sheet_name": table.sheet_name,
                    "subtable_id": table.subtable_id,
                    "row_index": row_index,
                    "col_range": _col_range(table.start_col, table.end_col),
                    "cell_range": _cell_range(row_index, table.start_col, table.end_col),
                    "headers": headers,
                    "fields": fields,
                    "warnings": [],
                },
            }
            chunks.append(doc)
        return chunks

    def _build_content(self, filename: str, table: _TableRegion, row_index: int, fields: dict[str, str]) -> str:
        source_file = Path(filename).name
        title = Path(filename).stem or source_file
        lines = [
            f"# {title} > {table.sheet_name} > 子表{table.subtable_id}",
            f"- 原文件: {source_file}",
            f"- 原Sheet: {table.sheet_name}",
            f"- 原位置: 行 {row_index}，列 {_col_range(table.start_col, table.end_col)}",
            "- 数据字段:",
        ]
        lines.extend(f"  * {key}: {value}" for key, value in fields.items())
        return "\n".join(lines)
