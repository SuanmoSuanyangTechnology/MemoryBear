"""Minimal Yuque client copied from the legacy authentication path."""

from __future__ import annotations

import asyncio
import json
import os
import re
import zlib
from collections.abc import Callable
from datetime import datetime
from typing import Any

import httpx
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .exceptions import YuqueAPIError, YuqueAuthError
from .models import YuqueDocInfo, YuqueRepoInfo


def _parse_time(value: Any) -> datetime | None:
    if not isinstance(value, str) or not value:
        return None
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError:
        return None


class YuqueAPIClient:
    def __init__(
        self,
        user_id: str,
        token: str,
        api_base_url: str = "https://www.yuque.com/api/v2",
        timeout: int = 30,
        client_factory: Callable[..., Any] = httpx.AsyncClient,
    ):
        self.user_id = user_id
        self._token = token
        self.api_base_url = api_base_url
        self.timeout = timeout
        self._client_factory = client_factory
        self._client: Any | None = None

    async def __aenter__(self) -> YuqueAPIClient:
        self._client = self._client_factory(
            base_url=self.api_base_url,
            timeout=self.timeout,
            headers={
                "Content-Type": "application/json",
                "X-Auth-Token": self._token,
                "User-Agent": "Yuque-Integration-Client",
            },
        )
        return self

    async def __aexit__(self, *_args: Any) -> None:
        if self._client is not None:
            await self._client.aclose()

    async def get_user_repos(self) -> list[YuqueRepoInfo]:
        if self._client is None:
            raise YuqueAPIError("HTTP client not initialized")
        try:
            response = await self._client.get(f"/users/{self.user_id}/repos")
        except httpx.HTTPError as exc:
            raise YuqueAPIError("Yuque network request failed") from exc
        if response.status_code == 401:
            raise YuqueAuthError("Yuque authentication failed", "401")
        if response.status_code != 200:
            raise YuqueAPIError("Yuque API request failed", str(response.status_code))
        body = response.json()
        result = []
        for item in body.get("data", []):
            if not isinstance(item, dict) or not isinstance(item.get("id"), int):
                continue
            result.append(
                YuqueRepoInfo(
                    id=item["id"],
                    type=str(item.get("type") or ""),
                    name=str(item.get("name") or ""),
                    namespace=str(item.get("namespace") or ""),
                    slug=str(item.get("slug") or ""),
                    description=item.get("description"),
                    public=int(item.get("public") or 0),
                    items_count=int(item.get("items_count") or 0),
                    created_at=_parse_time(item.get("created_at")),
                    updated_at=_parse_time(item.get("updated_at")),
                )
            )
        return result

    async def get_repo_docs(self, book_id: int) -> list[YuqueDocInfo]:
        """Return the published document snapshots exposed by one repository."""

        if self._client is None:
            raise YuqueAPIError("HTTP client not initialized")
        try:
            response = await self._client.get(f"/repos/{book_id}/docs")
        except httpx.HTTPError:
            raise YuqueAPIError("Yuque document list request failed") from None
        if response.status_code == 401:
            raise YuqueAuthError("Yuque authentication failed", "401")
        if response.status_code != 200:
            raise YuqueAPIError("Yuque document list request failed", str(response.status_code))
        documents = []
        for item in response.json().get("data", []):
            if not isinstance(item, dict):
                continue
            try:
                documents.append(self._document_from_payload(item, include_body=False))
            except (TypeError, ValueError):
                continue
        return documents

    async def get_doc_detail(self, document_id: int) -> YuqueDocInfo:
        """Load the raw document body required by the download path."""

        if self._client is None:
            raise YuqueAPIError("HTTP client not initialized")
        try:
            response = await self._client.get(
                f"/repos/docs/{document_id}",
                params={"raw": 1},
            )
        except httpx.HTTPError:
            raise YuqueAPIError("Yuque document detail request failed") from None
        if response.status_code == 401:
            raise YuqueAuthError("Yuque authentication failed", "401")
        if response.status_code != 200:
            raise YuqueAPIError("Yuque document detail request failed", str(response.status_code))
        payload = response.json().get("data") or {}
        if not isinstance(payload, dict):
            raise YuqueAPIError("Yuque document detail response is invalid")
        try:
            return self._document_from_payload(payload, include_body=True)
        except (TypeError, ValueError):
            raise YuqueAPIError("Yuque document detail response is invalid") from None

    async def download_document(self, doc: YuqueDocInfo, save_dir: str) -> str:
        """Download a Yuque document to the fixed legacy local representation."""

        try:
            if not doc.body:
                doc = await self.get_doc_detail(doc.id)
            filename = re.sub(r'[\\/:*?"<>|]', "_", doc.title)
            content = doc.body or ""
            extension = {
                "markdown": "md",
                "lake": "md",
                "html": "html",
                "lakesheet": "xlsx",
            }.get(doc.format, "txt")
            file_path = os.path.join(save_dir, f"{filename}.{extension}")
            await asyncio.to_thread(
                self._write_downloaded_document,
                doc.format,
                content,
                file_path,
            )
            return file_path
        except (YuqueAPIError, YuqueAuthError):
            raise
        except Exception:
            raise YuqueAPIError("Yuque document download failed") from None

    @staticmethod
    def _document_from_payload(
        payload: dict[str, Any],
        *,
        include_body: bool,
    ) -> YuqueDocInfo:
        created_at = _parse_time(payload.get("created_at"))
        updated_at = _parse_time(payload.get("updated_at"))
        if created_at is None or updated_at is None:
            raise ValueError("invalid Yuque document timestamps")
        return YuqueDocInfo(
            id=int(payload["id"]),
            type=str(payload.get("type") or ""),
            slug=str(payload.get("slug") or ""),
            title=str(payload.get("title") or ""),
            book_id=int(payload["book_id"]),
            format=str(payload.get("format") or "markdown"),
            body=str(payload.get("body") or "") if include_body else None,
            body_draft=payload.get("body_draft") if include_body else None,
            body_html=payload.get("body_html") if include_body else None,
            public=int(payload.get("public") or 0),
            status=int(payload.get("status") or 0),
            created_at=created_at,
            updated_at=updated_at,
            published_at=_parse_time(payload.get("published_at")),
            word_count=int(payload.get("word_count") or 0),
            cover=payload.get("cover"),
            description=payload.get("description"),
        )

    def _write_lakesheet(self, body: str, save_path: str) -> None:
        body_data = json.loads(body)
        compressed = bytes(body_data.get("sheet", ""), "latin-1")
        try:
            raw = zlib.decompress(compressed)
        except Exception:
            raise ValueError("Invalid or unsupported sheet data format") from None
        try:
            sheet_text = raw.decode("utf-8")
        except UnicodeDecodeError:
            sheet_text = raw.decode("gbk")
        sheets = json.loads(sheet_text)
        if not isinstance(sheets, list):
            raise ValueError("sheet data must be an array")

        workbook = Workbook()
        for sheet_index, sheet_data in enumerate(sheets):
            title = sheet_data.get("name", f"Sheet{sheet_index + 1}")
            worksheet = workbook.active if sheet_index == 0 else workbook.create_sheet()
            worksheet.title = title
            for column_index, style in enumerate(sheet_data.get("columns", []), start=1):
                worksheet.column_dimensions[get_column_letter(column_index)].width = (
                    style.get("size", 82.125) / 7.0
                )
            for row_index, style in enumerate(sheet_data.get("rows", []), start=1):
                worksheet.row_dimensions[row_index].height = style.get("size", 24) / 1.5
            for row_key, row in sheet_data.get("data", {}).items():
                for column_key, cell_data in row.items():
                    cell = worksheet.cell(row=int(row_key) + 1, column=int(column_key) + 1)
                    value = cell_data.get("value", "")
                    if isinstance(value, dict):
                        value = (
                            f"={value['formula']}"
                            if value.get("class") == "formula" and "formula" in value
                            else value.get("value", "")
                        )
                    cell.value = value
                    self._apply_cell_style(cell, cell_data.get("style", {}))
            for merge in sheet_data.get("mergeCells", {}).values():
                start_row = merge["row"] + 1
                start_column = merge["col"] + 1
                worksheet.merge_cells(
                    start_row=start_row,
                    start_column=start_column,
                    end_row=start_row + merge["rowCount"] - 1,
                    end_column=start_column + merge["colCount"] - 1,
                )
        workbook.save(save_path)

    def _write_downloaded_document(
        self,
        document_format: str,
        content: str,
        file_path: str,
    ) -> None:
        if document_format == "lakesheet":
            self._write_lakesheet(content, file_path)
            return
        if os.path.exists(file_path):
            os.remove(file_path)
        with open(file_path, "w", encoding="utf-8") as output:
            output.write(content)

    @staticmethod
    def _apply_cell_style(cell: Any, style: dict[str, Any]) -> None:
        horizontal = style.get("hAlign", "left")
        vertical = style.get("vAlign", "top")
        if horizontal not in {
            "general", "left", "center", "centerContinuous", "right", "fill",
            "justify", "distributed",
        }:
            horizontal = "left"
        if vertical not in {"top", "center", "justify", "distributed", "bottom"}:
            vertical = "top"
        color = YuqueAPIClient._convert_color(style.get("color", "#000000"))
        cell.font = Font(
            size=style.get("fontSize", 11),
            bold=style.get("fontWeight", False),
            italic=style.get("fontStyle", "normal") == "italic",
            underline="single" if style.get("underline", False) else None,
            color=color,
        )
        cell.alignment = Alignment(
            horizontal=horizontal,
            vertical=vertical,
            wrap_text=style.get("overflow") == "wrap",
        )
        background = YuqueAPIClient._convert_color(style.get("backColor"))
        if background:
            cell.fill = PatternFill(
                start_color=background,
                end_color=background,
                fill_type="solid",
            )

    @staticmethod
    def _convert_color(color: str | None) -> str | None:
        if not color:
            return None
        if color.startswith("#"):
            value = color[1:]
            return ("FF" + value).upper() if len(value) == 6 else value.upper()
        match = re.match(
            r"rgba?\(\s*(\d+)\s*,\s*(\d+)\s*,\s*(\d+)(?:\s*,\s*([\d.]+))?\s*\)",
            color,
        )
        if not match:
            return None
        red, green, blue = (int(match.group(index)) for index in range(1, 4))
        alpha = int(float(match.group(4) or 1) * 255)
        return f"{alpha:02X}{red:02X}{green:02X}{blue:02X}"


__all__ = ["YuqueAPIClient"]
